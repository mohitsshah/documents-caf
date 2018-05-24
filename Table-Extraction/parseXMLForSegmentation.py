# -*- coding: utf-8 -*-
"""
Created on Mon Apr 23 14:30:54 2018

@author: genesis
"""
import xml.etree.ElementTree
import cv2
import numpy as np
import math
import os
from openpyxl import Workbook


def findGroup(obj, num, dist, idx):
    keys = np.array(list(obj.keys()))
    if len(keys) > 0:
        distances = abs(keys - num)
        minval = np.min(distances)
        if minval <= dist:
            minidx = np.argmin(distances)
            obj[keys[minidx]]["count"] += 1
            obj[keys[minidx]]["indices"].append(idx)
        else:
            obj[num] = {
                "count": 1,
                "indices": [idx]
            }
    else:
        obj[num] = {
            "count": 1,
            "indices": [idx]
        }
    return obj


def trimSeg(seg):
    xc0 = 0
    yc0 = 0
    xc1, yc1 = seg.shape
    for i in range(xc1):
        line = seg[i, :]
        if np.any(line):
            xc0 = i
            break

    for i in range(yc1):
        line = seg[:, i]
        if np.any(line):
            yc0 = i
            break

    for i in range(xc1 - 1, xc0, -1):
        line = seg[i, :]
        if np.any(line):
            xc1 = i
            break

    for i in range(yc1 - 1, yc0, -1):
        line = seg[:, i]
        if np.any(line):
            yc1 = i
            break
    return seg[xc0:xc1, yc0:yc1]


def scoreHorzSeg(seg):
    height, width = seg.shape
    scores = []
    for i in range(height):
        line = seg[i, :]
        if np.any(line):
            scores.append(lineScore(line))
        else:
            pass
    return np.mean(scores)


def scoreVertSeg(seg):
    height, width = seg.shape

    scores = []
    for i in range(width):
        line = seg[:, i]
        if np.any(line):
            scores.append(lineScoreY(line))
        else:
            pass

    return np.mean(scores)


def getScoreNewY(obj, noComps, xory, xypoints, prtstr, seg_new_img, labels):
    score = []
    h, w = seg_new_img.shape
    keys = obj.keys()
    totalGroups = np.sum([obj[grp]["count"] for grp in obj])
    singleGroups = len([grp for grp in obj if obj[grp]["count"] == 1])

    if float(singleGroups) / totalGroups < 0.4:
        for k in keys:
            if obj[k]["count"] > 1:
                visGrp = np.zeros((h, w, 3), dtype=np.uint8)
                visSeg = np.zeros((h, w), dtype=np.uint8)
                indices = obj[k]["indices"]
                for i in indices:
                    visGrp[labels == i] = (0, 0, 255)
                    visSeg[labels == i] = 255
                visSeg = trimSeg(visSeg)

                score.append(scoreVertSeg(visSeg))

    if len(score) > 0:
        mn = np.mean(score)
    else:
        mn = 0.

    if not math.isnan(mn):
        return mn
    else:
        return 0.


def getScoreNew(obj, noComps, xory, xypoints, prtstr, seg_new_img, labels):
    score = []
    # Visualize Groups
    h, w = seg_new_img.shape
    visGrp = np.zeros((h, w, 3), dtype=np.uint8)
    keys = obj.keys()
    totalGroups = np.sum([obj[grp]["count"] for grp in obj])
    singleGroups = len([grp for grp in obj if obj[grp]["count"] == 1])

    if float(singleGroups) / totalGroups < 0.4:
        for k in keys:
            if obj[k]["count"] > 1:
                visGrp = np.zeros((h, w, 3), dtype=np.uint8)
                visSeg = np.zeros((h, w), dtype=np.uint8)
                indices = obj[k]["indices"]
                for i in indices:
                    visGrp[labels == i] = (0, 0, 255)
                    visSeg[labels == i] = 255
                visSeg = trimSeg(visSeg)

                score.append(scoreHorzSeg(visSeg))
    else:
        score.append(0.)
    if len(score) > 0:
        mn = np.mean(score)
    else:
        mn = 0.
    if not math.isnan(mn):
        return mn
    else:
        return 0.


def getGapScore(obj, xory, xypoints):
    ratios = []
    keys = obj.keys()
    for k in keys:
        gaps = []
        widths = []
        grpIdxs = obj[k]["indices"]
        grps = [(xypoints[i][xory], xypoints[i][xory + 2]) for i in range(len(xypoints)) if i in grpIdxs]
        lastDim = -1
        if len(grps) > 1:
            for grp in grps:
                widths.append(grp[1] - grp[0])
                if lastDim > -1:
                    gaps.append(grp[0] - lastDim)
                lastDim = grp[1]

            if abs(len(gaps) - len(widths)) < 3:
                totalgap = np.sum(gaps, dtype=np.float)
                totalwidth = np.sum(widths, dtype=np.float)
                ratio = totalgap / totalwidth
                ratios.append(ratio)

    score = np.mean(ratios)
    if score > 1:
        return 1.
    else:
        return score


def fetchXMLElements(src_xml_file):
    tree = xml.etree.ElementTree.parse(src_xml_file)
    page = list(tree.iter("page"))[0]
    all_textlines = list(tree.iter('textline'))
    all_figures = list(tree.iter('figure'))

    text_list = []

    for at in all_textlines:
        all_texts_for_lines = list(at.iter('text'))
        text = []
        flag = False
        for atfl in all_texts_for_lines:
            if atfl.text is not None and len(atfl.text) > 0 and atfl.text not in [" ", "\n"]:
                flag = True
                text.append(atfl.text)

        if set(text).issubset(set(['_', ' ', ''])):
            flag = False

        if flag:
            for atfl in all_texts_for_lines:
                if atfl.text is not None and len(atfl.text) > 0 and atfl.text not in [" ", "\n"]:
                    text_list.append([float(b) for b in atfl.get("bbox").split(",")])

    for fg in all_figures:
        all_texts_for_figures = list(fg.iter('text'))
        text = []
        flag = False
        for atff in all_texts_for_figures:
            if atff.text is not None and len(atff.text) > 0 and atff.text not in [" ", "\n"]:
                flag = True
                text.append(atff.text)

        if set(text).issubset(set(['_', ' ', ''])):
            flag = False

        if flag:
            for atff in all_texts_for_figures:
                if atff.text is not None and len(atff.text) > 0 and atff.text not in [" ", "\n"]:
                    text_list.append([float(b) for b in atff.get("bbox").split(",")])

    return page, text_list


def createNewPage(page, text_list):
    pagebb = [float(b) for b in page.get("bbox").split(",")]

    height = pagebb[3] - pagebb[1]
    width = pagebb[2] - pagebb[0]

    new_img = np.zeros((int(height), int(width)), dtype=np.uint8)
    for tl in text_list:
        bbox = tl
        x0 = int(round(height - bbox[3]))
        x1 = int(round(height - bbox[1] - 2))
        y0 = int(round(bbox[0]))
        y1 = int(round(bbox[2] - 2))
        new_img[x0:x1, y0:y1] = 255

    # remove small gaps between words
    ker = np.ones((1, int(round(0.01 * width))), dtype=np.uint8)
    new_img_dilated = cv2.dilate(new_img, ker, iterations=1)
    return new_img_dilated, height, width


def consecutive(data, stepsize=1):
    return np.split(data, np.where(np.diff(data) != stepsize)[0] + 1)


def lineScoreY(line):
    score = 0.
    counts = 0.
    if np.any(line):
        partitions = np.where(line[1:] != line[:-1])[0] + 1
        runs = []
        start = 0
        for p in partitions:
            runs.append((start, p - 1, line[start]))
            start = p
        runs.append((start, len(line) - 1, line[start]))
        if len(runs) >= 3:
            for i in range(len(runs)):
                if i + 2 < len(runs):
                    if runs[i][2] == 255:
                        counts += 1
                        fW = runs[i][1] - runs[i][0]
                        gW = runs[i + 1][1] - runs[i + 1][0]
                        tW = runs[i + 2][1] - runs[i + 2][0]

                        aW = float(fW + tW) / 2
                        ratio = float(gW) / (aW + gW)
                        if ratio > 0.05:
                            score += 1.
                        else:
                            pass
        else:
            score = 0.1
            counts = 1

    if score == 0.:
        return score
    else:
        return score / counts


def lineScore(line):
    score = 0.
    counts = 0.
    if np.any(line):
        partitions = np.where(line[1:] != line[:-1])[0] + 1
        runs = []
        start = 0
        for p in partitions:
            runs.append((start, p - 1, line[start]))
            start = p
        runs.append((start, len(line) - 1, line[start]))

        if len(runs) >= 3:
            for i in range(len(runs)):
                if i + 2 < len(runs):
                    if runs[i][2] == 255:
                        counts += 1
                        fW = runs[i][1] - runs[i][0]
                        gW = runs[i + 1][1] - runs[i + 1][0]
                        tW = runs[i + 2][1] - runs[i + 2][0]

                        aW = float(fW + tW) / 2
                        ratio = float(gW) / (aW + gW)
                        if ratio > 0.15:
                            score += 1.
        else:
            score = 0.1
            counts = 1

    if score == 0.:
        return score
    else:
        return score / counts


def scoreEachLine(new_img):
    height, width = new_img.shape
    scores = []
    for i in range(height):
        line = new_img[i, :]
        if np.any(line):
            scores.append(lineScore(line))
        else:
            scores.append(0.)

    score_new_img = np.zeros((height, width, 3), dtype=np.uint8)
    xpts, ypts = np.nonzero(new_img == 255)
    xypts = zip(xpts, ypts)
    for x, y in xypts:
        if scores[x] >= 0.66:
            score_new_img[x, y, :] = (0, 0, 255)
        elif scores[x] >= 0.33:
            score_new_img[x, y, :] = (0, 255, 0)
        else:
            score_new_img[x, y, :] = (255, 0, 0)
    return score_new_img


def cropimg(segimg):
    xc0, yc0 = 0, 0
    xc1, yc1 = segimg.shape
    for i in range(xc1):
        line = segimg[i, :]
        if np.any(line):
            xc0 = i
            break

    for i in range(yc1):
        line = segimg[:, i]
        if np.any(line):
            yc0 = i
            break

    for i in range(xc1 - 1, xc0, -1):
        line = segimg[i, :]
        if np.any(line):
            xc1 = i + 1
            break

    for i in range(yc1 - 1, yc0, -1):
        line = segimg[:, i]
        if np.any(line):
            yc1 = i + 1
            break

    return xc0, yc0, xc1, yc1


def cropline(line):
    return np.trim_zeros(line)


def segmentPage(new_img):
    # print("SEGMENTPAGE")
    xc0, yc0, xc1, yc1 = cropimg(new_img)
    cropped_img = new_img[xc0:xc1, yc0:yc1]

    height, width = cropped_img.shape
    min_gap = {}
    for i in range(width):
        line = cropped_img[:, i]
        if np.any(line):
            column = cropline(cropped_img[:, i])
            partitions = np.where(column[1:] != column[:-1])[0] + 1
            runs = []
            start = 0
            for p in partitions:
                runs.append((start, p - 1, column[start]))
                start = p
            runs.append((start, len(column) - 1, column[start]))
            for r in runs:
                if r[2] == 0:
                    if r[1] - r[0] in min_gap and r[1] - r[0] > 2:
                        min_gap[r[1] - r[0]] += 1
                    else:
                        min_gap[r[1] - r[0]] = 1

    if 0 in min_gap:
        min_gap.pop(0, None)

    gaps = list(min_gap.keys())
    gaps.sort()
    tc = np.sum([min_gap[i] for i in gaps])
    cdf = {}
    ccdf = 0

    for g in gaps:
        ccdf += min_gap[g]
        cdf[g] = float(ccdf) / tc
    # for g in gaps:
    #     print(g, min_gap[g], float(min_gap[g]) / tc, cdf[g])
    gapcounts = []

    for i in gaps:
        gapcounts.append(min_gap[i])

    min_gap_num = -1
    for i in range(1, len(gaps)):
        temp = gapcounts[:i]
        if temp:
            counts = np.sum(temp)
            if float(counts) / np.sum(gapcounts) > 0.35:
                min_gap_num = gaps[i - 1]
                min_gap_idx = i - 1
                break

    if min_gap_num == -1:
        return None

    # print("premingap", min_gap_num)
    min_gap_num = mingapLoop(min_gap_idx, gaps, min_gap, tc, min_gap_num)

    segments = []
    startSegment = -1
    endSegment = -1
    nullCount = 0
    # print("MIN_GAP:", np.max([min_gap_num]))
    for i in range(int(height)):
        line = cropped_img[i, :]
        if startSegment == -1:
            startSegment = i
        endSegment = i
        if not np.any(line):
            nullCount += 1
        else:
            nullCount = 0

        if nullCount > np.max([min_gap_num]):
            nullCount = 0
            segments.append((xc0 + startSegment, xc0 + endSegment))
            startSegment = -1
    if (xc0 + startSegment, xc0 + height + 1) not in segments and startSegment != -1:
        segments.append((xc0 + startSegment, xc0 + height + 1))
    return segments


def mingapLoop(min_gap_idx, gaps, min_gap, tc, min_gap_num):
    for i in range(min_gap_idx + 1, min_gap_idx + 1 + 5):
        if i < len(gaps) and gaps[i] - min_gap_num < 4:
            if float(min_gap[gaps[i]]) / tc > 0.1:
                min_gap_num = mingapLoop(i, gaps, min_gap, tc, min_gap_num)
                break
            elif float(min_gap[gaps[i]]) / tc > 0.01:
                min_gap_num = gaps[i] + 1
            else:
                min_gap_num += 2
                break
        else:
            min_gap_num += 2
            break
    return min_gap_num


def segmentPage3(new_img):
    # print("SEGMENTPAGE3")
    xc0, yc0, xc1, yc1 = cropimg(new_img)
    cropped_img = new_img[xc0:xc1, yc0:yc1]

    height, width = cropped_img.shape
    min_gap = {}
    for i in range(width):
        line = cropped_img[:, i]
        if np.any(line):
            column = cropline(cropped_img[:, i])
            partitions = np.where(column[1:] != column[:-1])[0] + 1
            runs = []
            start = 0
            for p in partitions:
                runs.append((start, p - 1, column[start]))
                start = p
            runs.append((start, len(column) - 1, column[start]))
            for r in runs:
                if r[2] == 0:
                    if r[1] - r[0] in min_gap and r[1] - r[0] > 2:
                        min_gap[r[1] - r[0]] += 1
                    else:
                        min_gap[r[1] - r[0]] = 1

    if 0 in min_gap:
        min_gap.pop(0, None)

    gaps = list(min_gap.keys())
    gaps.sort()
    tc = np.sum([min_gap[i] for i in gaps])
    cdf = {}
    ccdf = 0

    for g in gaps:
        ccdf += min_gap[g]
        cdf[g] = float(ccdf) / tc
    # for g in gaps:
    #     print(g, min_gap[g], float(min_gap[g]) / tc, cdf[g])
    gapcounts = []

    for i in gaps:
        gapcounts.append(min_gap[i])

    min_gap_num = -1
    for i in range(1, len(gaps)):
        temp = gapcounts[:i]
        if temp:
            counts = np.sum(temp)
            if float(counts) / np.sum(gapcounts) > 0.45:
                min_gap_num = gaps[i - 1]
                min_gap_idx = i - 1
                break

    if min_gap_num == -1:
        if not np.any(new_img):
            # print("RETURNED NONE")
            return None
        else:
            # print("RETURNED ORIGINAL")
            return [(xc0, xc1)]

    # print("premingap", min_gap_num)
    min_gap_num = mingapLoop(min_gap_idx, gaps, min_gap, tc, min_gap_num)

    segments = []
    startSegment = -1
    endSegment = -1
    nullCount = 0
    # print("MIN_GAP:", np.max([min_gap_num]))
    for i in range(int(height)):
        line = cropped_img[i, :]
        if startSegment == -1:
            startSegment = i
        endSegment = i
        if not np.any(line):
            nullCount += 1
        else:
            nullCount = 0

        if nullCount > np.max([min_gap_num]):
            nullCount = 0
            segments.append((xc0 + startSegment, xc0 + endSegment))
            startSegment = -1
    if (xc0 + startSegment, xc0 + height + 1) not in segments and startSegment != -1:
        segments.append((xc0 + startSegment, xc0 + height + 1))
    return segments


def segmentPage4(new_img):
    # print("SEGMENTPAGE4")
    xc0, yc0, xc1, yc1 = cropimg(new_img)
    cropped_img = new_img[xc0:xc1, yc0:yc1]

    height, width = cropped_img.shape
    min_gap = {}
    for i in range(width):
        line = cropped_img[:, i]
        if np.any(line):
            column = cropline(cropped_img[:, i])
            partitions = np.where(column[1:] != column[:-1])[0] + 1
            runs = []
            start = 0
            for p in partitions:
                runs.append((start, p - 1, column[start]))
                start = p
            runs.append((start, len(column) - 1, column[start]))
            for r in runs:
                if r[2] == 0:
                    if r[1] - r[0] in min_gap and r[1] - r[0] > 2:
                        min_gap[r[1] - r[0]] += 1
                    else:
                        min_gap[r[1] - r[0]] = 1

    if 0 in min_gap:
        min_gap.pop(0, None)

    gaps = list(min_gap.keys())
    gaps.sort()
    tc = np.sum([min_gap[i] for i in gaps])
    cdf = {}
    ccdf = 0

    for g in gaps:
        ccdf += min_gap[g]
        cdf[g] = float(ccdf) / tc
    # for g in gaps:
    #     print(g, min_gap[g], float(min_gap[g]) / tc, cdf[g])
    gapcounts = []

    for i in gaps:
        gapcounts.append(min_gap[i])

    min_gap_num = -1
    for i in range(1, len(gaps)):
        temp = gapcounts[:i]
        if temp:
            counts = np.sum(temp)
            if float(counts) / np.sum(gapcounts) > 0.45:
                min_gap_num = gaps[i - 1]
                break

    if min_gap_num == -1:
        if not np.any(new_img):
            # print("RETURNED NONE")
            return None
        else:
            # print("RETURNED ORIGINAL")
            return [(xc0, xc1)]

    # print("premingap", min_gap_num)

    segments = []
    startSegment = -1
    endSegment = -1
    nullCount = 0
    # print("MIN_GAP:", np.max([min_gap_num]))
    for i in range(int(height)):
        line = cropped_img[i, :]
        if startSegment == -1:
            startSegment = i
        endSegment = i
        if not np.any(line):
            nullCount += 1
        else:
            nullCount = 0

        if nullCount > np.max([min_gap_num]):
            nullCount = 0
            segments.append((xc0 + startSegment, xc0 + endSegment))
            startSegment = -1
    if (xc0 + startSegment, xc0 + height + 1) not in segments and startSegment != -1:
        segments.append((xc0 + startSegment, xc0 + height + 1))
    return segments


def partitionPage(new_img, page_height):
    xc0, yc0, xc1, yc1 = cropimg(new_img)
    cropped_img = new_img[xc0:xc1, yc0:yc1]
    height, width = cropped_img.shape
    counts = []
    for i in range(height):
        line = cropped_img[i, :]
        if not np.any(line):
            counts.append(1)
        else:
            counts.append(0)
    counts = np.array(counts)

    partitions = np.where(counts[1:] != counts[:-1])[0] + 1
    runs = []
    start = 0
    for p in partitions:
        if counts[start] == 1:
            if (p - 1 - start) > 0.03 * height:
                runs.append((xc0 + start, xc0 + p))
        start = p
    if counts[start] == 1 and (xc0 + start, xc0 + height) not in runs:
        runs.append((xc0 + start, xc0 + height))

    if len(runs) == 0:
        return [(new_img, None)]
    breaks = []
    for r in runs:
        breaks.append((float(r[0]) + float(r[1])) / 2)
    pageParts = []
    start = 0
    for b in breaks:
        pageParts.append(
            (new_img[start:int(np.floor(b)), :], None, start, int(np.floor(b))))
        start = int(np.floor(b))
    pageParts.append((new_img[start:, :], None, start, int(page_height)))
    return pageParts


def segmentPage2(new_img):
    #    OLD APPROACH#
    height, width = new_img.shape
    min_gap = []
    for i in range(width):
        column = new_img[:, i]
        partitions = np.where(column[1:] != column[:-1])[0] + 1
        runs = []
        start = 0
        for p in partitions:
            runs.append((start, p - 1, column[start]))
            start = p
        runs.append((start, len(column) - 1, column[start]))
        for r in runs:
            if r[2] == 0:
                min_gap.append(r[1] - r[0])

    min_gap = list(set(min_gap))
    min_gap.sort()
    if len(min_gap) >= 5:
        min_gap_num = int(math.ceil(np.mean(min_gap[0:5])))
    else:
        min_gap_num = int(math.ceil(np.mean(min_gap[0:len(min_gap)])))

    segments = []
    startSegment = -1
    endSegment = -1
    nullCount = 0
    min_perc_gap = int(round(0.01 * height))
    for i in range(int(height)):
        line = new_img[i, :]
        if startSegment == -1:
            startSegment = i
        endSegment = i
        if not np.any(line):
            nullCount += 1
        else:
            nullCount = 0

        if nullCount > np.max([min_gap_num, min_perc_gap]):
            nullCount = 0
            segments.append((startSegment, endSegment + 1))
            startSegment = -1

    if (startSegment, height + 1) not in segments and startSegment != -1:
        segments.append((startSegment, height + 1))

    return segments


def processSegmentsForTables(segs, img, new_img):
    seg_img = img[segs[0]:segs[1], :]

    seg_new_img = new_img[segs[0]:segs[1], :]
    xc0 = 0
    yc0 = 0
    xc1, yc1 = seg_new_img.shape
    for i in range(xc1):
        line = seg_new_img[i, :]
        if np.any(line):
            xc0 = i
            break

    for i in range(yc1):
        line = seg_new_img[:, i]
        if np.any(line):
            yc0 = i
            break

    for i in range(xc1 - 1, xc0, -1):
        line = seg_new_img[i, :]
        if np.any(line):
            xc1 = i
            break

    for i in range(yc1 - 1, yc0, -1):
        line = seg_new_img[:, i]
        if np.any(line):
            yc1 = i
            break

    seg_img = seg_img[xc0:xc1, yc0:yc1]
    seg_new_img = seg_new_img[xc0:xc1, yc0:yc1]

    if np.average(seg_new_img) / 255 > 0.10 and np.average(seg_new_img) / 255 < 0.75:
        # print(np.average(seg_new_img) / 255)
        unique, counts = np.unique(seg_new_img, return_counts=True)
        cdict = dict(zip(unique, counts))
        if 255 in cdict and 0 in cdict:
            output = cv2.connectedComponentsWithStats(seg_new_img, 4, cv2.CV_32S)
            labels = output[1]
            stats = output[2]
            xypoints = []
            for i in range(len(stats)):
                y0 = stats[i][0]
                x0 = stats[i][1]
                y1 = stats[i][0] + stats[i][2] - 1
                x1 = stats[i][1] + stats[i][3] - 1
                xypoints.append((x0, y0, x1, y1))
            centroids = output[3]
            xcentroids = centroids[:, 1]
            ycentroids = centroids[:, 0]

            noOfComponents = len(stats)

            NZComponentIndices = []
            for i in range(noOfComponents):
                xpts, ypts = np.nonzero(output[1] == i)
                if seg_new_img[xpts[0], ypts[0]] != 0:
                    NZComponentIndices.append(i)

            noOfNZComponents = len(NZComponentIndices)

            xminsg = {}
            yminsg = {}
            xmaxsg = {}
            ymaxsg = {}
            xcentroidsg = {}
            ycentroidsg = {}
            for i in NZComponentIndices:
                xminsg = findGroup(xminsg, xypoints[i][0], 2, i)
                yminsg = findGroup(yminsg, xypoints[i][1], 2, i)
                xmaxsg = findGroup(xmaxsg, xypoints[i][2], 2, i)
                ymaxsg = findGroup(ymaxsg, xypoints[i][3], 2, i)
                xcentroidsg = findGroup(xcentroidsg, xcentroids[i], 2, i)
                ycentroidsg = findGroup(ycentroidsg, ycentroids[i], 2, i)

            score = []

            sc_xminsg = getScoreNew(xminsg, noOfNZComponents, 1, xypoints, "xminsg", seg_new_img, labels) or 0.
            sc_xmaxsg = getScoreNew(xmaxsg, noOfNZComponents, 1, xypoints, "xmaxsg", seg_new_img, labels) or 0.
            sc_xcentroidsg = getScoreNew(xcentroidsg, noOfNZComponents, 1, xypoints, "xcentroid", seg_new_img,
                                         labels) or 0.
            sc_yminsg = getScoreNewY(yminsg, noOfNZComponents, 0, xypoints, "yminsg", seg_new_img, labels) or 0.
            sc_ymaxsg = getScoreNewY(ymaxsg, noOfNZComponents, 0, xypoints, "ymaxsg", seg_new_img, labels) or 0.
            sc_ycentroidsg = getScoreNewY(ycentroidsg, noOfNZComponents, 0, xypoints, "ycentroid", seg_new_img,
                                          labels) or 0.

            score.append(np.max([sc_xminsg, sc_xmaxsg, sc_xcentroidsg]))
            score.append(np.max([sc_yminsg, sc_ymaxsg, sc_ycentroidsg]))

            # print(score)
            return score, seg_img, seg_new_img, xc0, yc0, xc1, yc1
        return [0., 0.], "", "", "", "", "", ""
    return [0., 0.], "", "", "", "", "", ""


def processSegmentsForTables2(seg, img, new_img):
    # seg_img = img[seg[0]:seg[2], seg[1]:seg[3]]
    seg_img = None
    seg_new_img = new_img[seg[0]:seg[2], seg[1]:seg[3]]
    #    cv2.imshow("seg",seg_new_img)
    #    cv2.imshow("seg_img",seg_img)
    #    cv2.waitKey()
    if np.average(seg_new_img) / 255 > 0.10 and np.average(seg_new_img) / 255 < 0.75:
        unique, counts = np.unique(seg_new_img, return_counts=True)
        cdict = dict(zip(unique, counts))
        if 255 in cdict and 0 in cdict:
            output = cv2.connectedComponentsWithStats(seg_new_img, 4, cv2.CV_32S)
            labels = output[1]
            stats = output[2]
            xypoints = []
            for i in range(len(stats)):
                y0 = stats[i][0]
                x0 = stats[i][1]
                y1 = stats[i][0] + stats[i][2] - 1
                x1 = stats[i][1] + stats[i][3] - 1
                xypoints.append((x0, y0, x1, y1))
            centroids = output[3]
            xcentroids = centroids[:, 1]
            ycentroids = centroids[:, 0]

            noOfComponents = len(stats)

            NZComponentIndices = []
            for i in range(noOfComponents):
                xpts, ypts = np.nonzero(output[1] == i)
                if seg_new_img[xpts[0], ypts[0]] != 0:
                    NZComponentIndices.append(i)

            noOfNZComponents = len(NZComponentIndices)

            xminsg = {}
            yminsg = {}
            xmaxsg = {}
            ymaxsg = {}
            xcentroidsg = {}
            ycentroidsg = {}
            for i in NZComponentIndices:
                xminsg = findGroup(xminsg, xypoints[i][0], 2, i)
                yminsg = findGroup(yminsg, xypoints[i][1], 2, i)
                xmaxsg = findGroup(xmaxsg, xypoints[i][2], 2, i)
                ymaxsg = findGroup(ymaxsg, xypoints[i][3], 2, i)
                xcentroidsg = findGroup(xcentroidsg, xcentroids[i], 2, i)
                ycentroidsg = findGroup(ycentroidsg, ycentroids[i], 2, i)

            score = []

            sc_xminsg = getScoreNew(xminsg, noOfNZComponents, 1, xypoints, "xminsg", seg_new_img, labels) or 0.
            sc_xmaxsg = getScoreNew(xmaxsg, noOfNZComponents, 1, xypoints, "xmaxsg", seg_new_img, labels) or 0.
            sc_xcentroidsg = getScoreNew(xcentroidsg, noOfNZComponents, 1, xypoints, "xcentroid", seg_new_img,
                                         labels) or 0.
            sc_yminsg = getScoreNewY(yminsg, noOfNZComponents, 0, xypoints, "yminsg", seg_new_img, labels) or 0.
            sc_ymaxsg = getScoreNewY(ymaxsg, noOfNZComponents, 0, xypoints, "ymaxsg", seg_new_img, labels) or 0.
            sc_ycentroidsg = getScoreNewY(ycentroidsg, noOfNZComponents, 0, xypoints, "ycentroid", seg_new_img,
                                          labels) or 0.

            # print(sc_xminsg, sc_xmaxsg, sc_xcentroidsg, sc_yminsg, sc_ymaxsg, sc_ycentroidsg)
            score.append(np.max([sc_xminsg, sc_xmaxsg, sc_xcentroidsg]))
            score.append(np.max([sc_yminsg, sc_ymaxsg, sc_ycentroidsg]))

            return score, seg_img, seg_new_img, seg[0], seg[1], seg[2], seg[3]
        return [0., 0.], "", "", "", "", "", ""
    return [0., 0.], "", "", "", "", "", ""


def getRatio(arr):
    nz = float(np.count_nonzero(arr))
    return nz / len(arr)


def extractTables(new_img, potentialTables, start, stop):
    tables = []
    for pt in potentialTables:
        potTable = new_img[pt[0]:pt[2], pt[1]:pt[3]]
        height, width = potTable.shape
        xmin = -1
        xmax = -1
        for i in range(height):
            line = potTable[i, :]
            ratio = getRatio(line)
            if ratio < 1:
                xmin = i
                break

        for i in range(height - 1, 0, -1):
            line = potTable[i, :]
            ratio = getRatio(line)
            if ratio < 0.8:
                xmax = i
                break
        potTable = potTable[xmin:xmax, :]
        # cv2.rectangle(main_img, (pt[1], start + pt[0] + xmin), (pt[3], start + pt[2] - (height - xmax - 2)),
        #               (0, 255, 0))
        #        tables.append((main_img[start+pt[0]+xmin:start+pt[2]-(height-xmax-2),pt[1]:pt[3]],potTable))
        tables.append([start + pt[0] + xmin, start + pt[2] - (height - xmax - 2), pt[1], pt[3], potTable])

    return tables


#    cv2.imshow("img",img)
#    cv2.imshow("new_img",new_img)
#    cv2.waitKey()    
#    cv2.destroyAllWindows()

def getPotentialTables(segments, img, new_img):
    potentialTables = []
    for segs in segments:
        score, seg_img, seg_new_img, xc0, yc0, xc1, yc1 = processSegmentsForTables(segs, img, new_img)
        avgscore = np.mean(score)
        xscore = score[0]
        yscore = score[1]
        if avgscore > 0.45 and xscore > 0.05 and yscore > 0.05:
            potentialTables.append([segs[0] + xc0, yc0 - 1, segs[0] + xc1, yc1 - 1])
    return new_img, potentialTables


def getPotentialTablesFromSegs(segs, img, new_img):
    potentialTables = []
    for seg in segs:
        score, seg_img, seg_new_img, xc0, yc0, xc1, yc1 = processSegmentsForTables2(seg, img, new_img)
        avgscore = np.mean(score)
        xscore = score[0]
        yscore = score[1]
        if avgscore > 0.45 and xscore > 0.05 and yscore > 0.05:
            potentialTables.append([seg[0], seg[1], seg[2], seg[3]])
    return new_img, potentialTables


def rectAinrectB(xa0, xa1, ya0, ya1, xb0, xb1, yb0, yb1):
    if xa0 >= xb0 and xa0 < xb1:
        if xa1 >= xb0 and xa1 <= xb1:
            if ya0 >= yb0 and ya0 < yb1:
                if ya1 >= yb0 and ya1 <= yb1:
                    return True
                else:
                    return False
            else:
                return False
        else:
            return False
    else:
        return False


def fetchXMLElementsFromRegion(src_xml_file, x0, x1, y0, y1):
    tree = xml.etree.ElementTree.parse(src_xml_file)
    page = list(tree.iter("page"))[0]
    pagebb = [float(b) for b in page.get("bbox").split(",")]
    height = pagebb[3] - pagebb[1]
    all_textlines = list(tree.iter('textline'))
    all_figures = list(tree.iter('figure'))

    text_list = []

    for at in all_textlines:
        bbox = at.get("bbox").split(",")
        all_texts_for_lines = list(at.iter('text'))
        text = []
        flag = False
        for atfl in all_texts_for_lines:
            if atfl.text is not None and len(atfl.text) > 0 and atfl.text not in ["\n"]:
                flag = True
                text.append(atfl.text)

        if set(text).issubset(set(['_', ' ', ''])):
            flag = False

        if flag:
            for atfl in all_texts_for_lines:
                if atfl.text is not None and atfl.get("bbox") and len(atfl.text) > 0 and atfl.text not in ["\n"]:
                    bbox = [float(a) for a in atfl.get("bbox").split(",")]
                    xc0 = int(round(height - bbox[3]))
                    xc1 = int(round(height - bbox[1] - 2))
                    yc0 = int(round(bbox[0]))
                    yc1 = int(round(bbox[2] - 2))
                    if rectAinrectB(xc0, xc1, yc0, yc1, x0, x1, y0, y1):
                        text_list.append([xc0, xc1, yc0, yc1, atfl.text])

    for fg in all_figures:
        all_texts_for_figures = list(fg.iter('text'))
        text = []
        flag = False
        for atff in all_texts_for_figures:
            if atff.text is not None and len(atff.text) > 0 and atff.text not in ["\n"]:
                flag = True
                text.append(atff.text)

        if set(text).issubset(set(['_', ' ', ''])):
            flag = False

        if flag:
            for atff in all_texts_for_figures:
                if atff.text is not None and atff.get("bbox") and len(atff.text) > 0 and atff.text not in ["\n"]:
                    bbox = atff.get("bbox").split(",")
                    xc0 = int(round(height - bbox[3]))
                    xc1 = int(round(height - bbox[1] - 2))
                    yc0 = int(round(bbox[0]))
                    yc1 = int(round(bbox[2] - 2))
                    if rectAinrectB(xc0, xc1, yc0, yc1, x0, x1, y0, y1):
                        text_list.append([xc0, xc1, yc0, yc1, atff.text])

    return page, text_list


def createNewRegion(page, text_list):
    pagebb = [float(b) for b in page.get("bbox").split(",")]

    height = pagebb[3] - pagebb[1]
    width = pagebb[2] - pagebb[0]

    new_img = np.zeros((int(height), int(width)), dtype=np.uint8)

    for tl in text_list:
        new_img[tl[0]:tl[1] + 1, tl[2]:tl[3]] = 255

    # remove small gaps between words
    ker = np.ones((1, int(round(0.01 * width))), dtype=np.uint8)
    new_img = cv2.dilate(new_img, ker, iterations=2)
    return new_img, height, width


def fetchText(textList):
    currentX = textList[0][0]
    text = ""
    for tl in textList:
        if tl[0] != currentX:
            currentX = tl[0]
            text += "\n"
        text += tl[4]
    return text


def fetchWordsFromRegion(words, y0, y1, x0, x1):
    print (x0, y0, x1, y1)
    word_list = []
    for word in words:
        xc0 = word[0]
        xc1 = word[2]
        yc0 = word[1]
        yc1 = word[3]
        # print (word)
        if np.abs(xc0 - x0) < 10 and np.abs(xc1 - x1) < 10 and np.abs(yc0 - y0) < 10 and np.abs(yc1 - y1) < 10:
            word_list.append([yc0, yc1, xc0, xc1, word[-1]])
    return word_list

def getRows2(seg_img, words, x0, x1, y0, y1):
    h, w = seg_img.shape
    xcounts = np.zeros(h, dtype=np.uint8)
    for i in range(h):
        row = seg_img[i, :]
        xcounts[i] = np.sum(row)
    xpartitions = np.where(xcounts[1:] != xcounts[:-1])[0] + 1
    xruns = []
    xstart = 0
    for p in xpartitions:
        xruns.append((xstart, p - 1, xcounts[xstart] > 0))
        xstart = p
    xruns.append((xstart, len(xcounts) - 1, xcounts[xstart] > 0))
    xbreaks = []
    for r in xruns:
        if r[2] == False:
            xbreaks.append(int(float(r[0] + r[1]) / 2))

    ycounts = np.zeros(w, dtype=np.uint8)
    for i in range(w):
        col = seg_img[:, i]
        ysum = np.sum(col) / 255
        if ysum < 0.01:
            ycounts[i] = 1
    ypartitions = np.where(ycounts[1:] != ycounts[:-1])[0] + 1
    yruns = []
    ystart = 0
    for p in ypartitions:
        yruns.append((ystart, p - 1, ycounts[ystart] > 0))
        ystart = p

    yruns.append((xstart, len(ycounts) - 1, ycounts[ystart] > 0))
    ybreaks = []
    for r in yruns:
        if r[2] == True:
            ybreaks.append(int(float(r[0] + r[1]) / 2))
    ybreaks.append(w + 1)
    xbreaks.append(h + 1)

    xstart = 0
    wb = Workbook()
    ws = wb.active
    for xb in xbreaks:
        xend = xb
        exRow = seg_img[xstart:xend, :]
        output = cv2.connectedComponentsWithStats(exRow, 4, cv2.CV_32S)
        noComponents = output[0]
        compImg = output[1]
        cellDict = {i: [] for i in range(len(ybreaks))}
        compText = {}
        for cidx in range(noComponents):
            # print (cidx, compImg.shape)
            comp = np.where(compImg == cidx)
            # print (comp)
            points = list(zip(comp[0], comp[1]))
            if exRow[points[0]] == 0:
                continue
            else:
                minY = np.min(comp[1])
                maxY = np.max(comp[1])
                # textList = \
                #     fetchXMLElementsFromRegion(src_xml_file, x0 + xstart - 1, x0 + xend + 1, y0 + minY - 1,
                #                                y0 + maxY + 1)[
                #         1]
                textList = fetchWordsFromRegion(words, x0 + xstart - 1, x0 + xend + 1, y0 + minY - 1, y0 + maxY + 1)
                print (textList)
                if len(textList) > 0:
                    textList = sorted(textList, key=lambda x: (x[0], x[2]))
                    text = fetchText(textList)
                    text = text.lstrip("'").lstrip()
                    compText[cidx] = text
                    for i in range(len(ybreaks)):
                        if i == 0:
                            ystrt = 0
                        else:
                            ystrt = ybreaks[i - 1]
                        ystp = ybreaks[i]

                        if (minY < ystrt and maxY > ystrt) or (minY < ystp and maxY > ystp) or (
                                minY >= ystrt and maxY < ystp):
                            cellDict[i].append(cidx)

        row = []
        for i in cellDict:
            celltext = ""
            for j in cellDict[i]:
                celltext += " " + compText[j]
            row.append(celltext)
        mergeCells = []
        compCells = {}
        for i in compCells:
            row[compCells[i][0]] = compText[i]
            if len(compCells[i]) > 1:
                mergeCells.append([j for j in compCells[i]])

        ws.append(row)
        if len(mergeCells) > 0:
            for m in mergeCells:
                mm = sorted(m)
                ws.merge_cells(start_row=ws.max_row, start_column=mm[0], end_row=ws.max_row, end_column=mm[-1])

        xstart = xb
    fname = str(x0) + str(x1) + str(y0) + str(y1)
    if not os.path.isdir("tables"):
        os.mkdir("tables")
    wb.save("tables/" + fname + ".xlsx")


def extractData(tables, words):
    for t in tables:
        getRows2(t[4], words, t[0], t[1], t[2], t[3])
